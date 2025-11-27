import re
import os
import sys
import smtplib
import ssl
import subprocess
import threading
import socket
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Sequence, cast
from xml.sax.saxutils import escape
import textwrap

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

GROUP_HEADER_PATTERN = re.compile(r"^(havo|vwo)\s*\d", re.IGNORECASE)
LEFT_MARGIN = 90
RIGHT_MARGIN = 90
TOP_MARGIN = 80
BOTTOM_MARGIN = 72
TABLE_FIRST_COLUMN_WIDTH = 190
# Standaard map voor openen van Excel-bestanden
if sys.platform.startswith("win") or os.name == "nt":
    DEFAULT_INITIAL_DIR = Path(r"H:\Gedeelde drives\Informatica")
else:
    DEFAULT_INITIAL_DIR = Path(
        "/Users/servaaswinder/Library/CloudStorage/GoogleDrive-servaas.winder@northgo-college.nl/Gedeelde drives/Informatica"
    )
SENDER_EMAIL = "servaas.winder@northgo-college.nl"
SENDER_PASSWORD = "gxtq rish aril ockn"
COMMENT_LABELS = {"opmerking", "opmerkingen", "comments", "notities"}

STYLES = getSampleStyleSheet()
STYLE_ASSIGNMENT = ParagraphStyle(
    "AssignmentHeading",
    parent=STYLES["Heading1"],
    fontName="Helvetica-Bold",
    fontSize=16,
    leading=20,
    spaceAfter=12,
)
STYLE_INFO = ParagraphStyle(
    "Info",
    parent=STYLES["Normal"],
    fontSize=11,
    leading=14,
    spaceAfter=4,
)
STYLE_SECTION = ParagraphStyle(
    "SectionHeading",
    parent=STYLES["Heading4"],
    fontName="Helvetica-Bold",
    fontSize=12,
    leading=15,
    spaceBefore=10,
    spaceAfter=6,
)
STYLE_TABLE_HEADER = ParagraphStyle(
    "TableHeader",
    parent=STYLES["Normal"],
    fontName="Helvetica-Bold",
    fontSize=11,
    leading=14,
)
STYLE_TABLE_BODY = ParagraphStyle(
    "TableBody",
    parent=STYLES["Normal"],
    fontSize=11,
    leading=14,
)


def _cell_to_text(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    text = str(cell).strip()
    if text == ".":
        return ""
    return text


def _sanitize_text(value: str) -> str:
    cleaned = value.replace("\r", " ").replace("\n", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        cleaned = "-"
    return cleaned.encode("latin-1", "replace").decode("latin-1")


def _format_text(value: object) -> str:
    if value is None:
        return "-"
    text = str(value).strip()
    if not text:
        return "-"
    return _sanitize_text(text)


def _paragraph_text(value: object) -> str:
    return escape(_format_text(value), {"'": "&apos;", '"': "&quot;"})


def _safe_row_value(row: Sequence[str], index: int) -> str:
    return row[index] if index < len(row) else ""


def _extract_http_link(row_cells: Sequence[Any], row_values: Sequence[str]) -> str:
    for cell in row_cells:
        hyperlink = getattr(cell, "hyperlink", None)
        if not hyperlink or not getattr(hyperlink, "target", None):
            continue
        target = hyperlink.target.strip()
        if target.lower().startswith("http"):
            return target
    return next((value for value in row_values if value.lower().startswith("http")), "")


def split_assignment_columns(columns: list[tuple[int, str]]) -> tuple[list[tuple[int, str]], tuple[int, str] | None]:
    score_columns: list[tuple[int, str]] = []
    comment_column: tuple[int, str] | None = None

    for col_index, column_name in columns:
        normalized = column_name.strip().lower()
        if normalized in COMMENT_LABELS and comment_column is None:
            comment_column = (col_index, column_name)
        else:
            score_columns.append((col_index, column_name))

    if not score_columns:
        score_columns = columns[:]
    return score_columns, comment_column


def build_ascii_table(rows: list[tuple[str, str]]) -> list[str]:
    if not rows:
        return []

    label_width = max(len("Categorie"), max(len(_format_text(label)) for label, _ in rows))

    wrapped_rows: list[list[str]] = []
    value_width = len("Waarde")
    wrap_width = 60
    for label, value in rows:
        wrapped = textwrap.wrap(_format_text(value), wrap_width) or ["-"]
        wrapped_rows.append(wrapped)
        value_width = max(value_width, max(len(line) for line in wrapped))

    top_border = f"+{'-' * (label_width + 2)}+{'-' * (value_width + 2)}+"
    header = f"| {'Categorie'.ljust(label_width)} | {'Waarde'.ljust(value_width)} |"
    header_sep = f"+{'=' * (label_width + 2)}+{'=' * (value_width + 2)}+"
    lines = [top_border, header, header_sep]

    for (label, _), wrapped in zip(rows, wrapped_rows):
        label_text = _format_text(label)
        for idx, line in enumerate(wrapped):
            label_part = label_text if idx == 0 else ""
            lines.append(f"| {label_part.ljust(label_width)} | {line.ljust(value_width)} |")
        lines.append(top_border)

    return lines


def compose_student_summary(
    assignment: str,
    student: dict,
    score_columns: list[tuple[int, str]],
    comment_column: tuple[int, str] | None,
) -> str:
    lines: list[str] = [f"Opdracht: {assignment}"]
    if student.get("group"):
        lines.append(f"Groep: {student['group']}")
    website = student.get("website")
    if website:
        lines.append(f"Website: {website}")
    lines.append("")
    lines.append("Scores:")
    rows = [
        (column_name, _safe_row_value(student["row"], col_index))
        for col_index, column_name in score_columns
    ]
    lines.extend(build_ascii_table(rows))
    if comment_column:
        col_index, label = comment_column
        comment_value = _safe_row_value(student["row"], col_index)
        if comment_value:
            lines.append("")
            lines.append(f"{label}:")
            lines.append(_format_text(comment_value))
    return "\n".join(lines)


def _build_email(subject: str, body: str, recipients: list[str]) -> EmailMessage:
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SENDER_EMAIL
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)
    return msg


def _login_credentials(smtp: smtplib.SMTP) -> None:
    password = SENDER_PASSWORD.replace(" ", "")
    smtp.login(SENDER_EMAIL, password)


def send_email_via_gmail(subject: str, body: str, recipients: list[str]) -> None:
    if not recipients:
        raise ValueError("Geen ontvangers opgegeven.")

    msg = _build_email(subject, body, recipients)
    context = ssl.create_default_context()
    # Probeer eerst de gebruikelijke SSL-poort; val terug op STARTTLS als de verbinding blokkeert
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context, timeout=20) as smtp:
            _login_credentials(smtp)
            smtp.send_message(msg)
            return
    except (socket.timeout, OSError):
        pass

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as smtp:
        smtp.starttls(context=context)
        _login_credentials(smtp)
        smtp.send_message(msg)


def build_email_body(
    assignment: str,
    student: dict,
    score_columns: list[tuple[int, str]],
    comment_column: tuple[int, str] | None,
) -> str:
    greeting_name = student.get("name") or "leerling"
    lines: list[str] = [
        f"Beste {greeting_name},",
        "",
        f"In deze e-mail vind je jouw resultaten voor {assignment}.",
        "",
    ]
    lines.append(compose_student_summary(assignment, student, score_columns, comment_column))
    lines.append("")
    lines.append("Met vriendelijke groet,")
    lines.append("Servaas Winder")
    return "\n".join(lines)


def parse_excel(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True)
    # Hint type checker that "active" is a Worksheet (not Optional)
    sheet: Worksheet = cast(Worksheet, workbook.active)
    try:
        row_data: list[tuple[list[str], str]] = []
        for row_cells in sheet.iter_rows():
            row_values = [_cell_to_text(cell.value) for cell in row_cells]
            website_link = _extract_http_link(row_cells, row_values)
            row_data.append((row_values, website_link))
    finally:
        workbook.close()

    if len(row_data) < 2:
        raise RuntimeError("Het Excel-bestand bevat te weinig rijen om de koppen te lezen.")

    iterator = iter(row_data)
    assignment_row, _ = next(iterator)
    column_row, _ = next(iterator)

    assignment_headers: list[str] = []
    assignment_order: list[str] = []
    current_assignment = ""
    for cell in assignment_row:
        cell = cell.strip()
        if cell:
            current_assignment = cell
            if current_assignment not in assignment_order:
                assignment_order.append(current_assignment)
        assignment_headers.append(current_assignment)

    column_headers = [cell.strip() for cell in column_row]

    assignment_columns: dict[str, list[tuple[int, str]]] = {}
    for idx, assignment in enumerate(assignment_headers):
        assignment = assignment.strip()
        column_name = (
            column_headers[idx]
            if idx < len(column_headers) and column_headers[idx]
            else f"Kolom {idx + 1}"
        )
        if assignment:
            assignment_columns.setdefault(assignment, []).append((idx, column_name))

    group_counts: list[tuple[str, int]] = []
    current_group_index: int | None = None
    student_count = 0
    students: list[dict] = []

    for row, website_link in iterator:
        stripped_row = [cell.strip() for cell in row]
        if not any(stripped_row):
            continue

        name = stripped_row[0]
        rest_has_data = any(stripped_row[1:])

        if name and not rest_has_data and GROUP_HEADER_PATTERN.match(name):
            group_counts.append((name, 0))
            current_group_index = len(group_counts) - 1
            continue

        if not name:
            continue

        student_count += 1
        if current_group_index is not None:
            group_name, count = group_counts[current_group_index]
            group_counts[current_group_index] = (group_name, count + 1)
            group_for_student = group_name
        else:
            if group_counts and group_counts[0][0] == "(onbekende groep)":
                group_name, count = group_counts[0]
                group_counts[0] = (group_name, count + 1)
            else:
                group_counts.insert(0, ("(onbekende groep)", 1))
            group_for_student = group_counts[0][0]
            current_group_index = 0

        students.append({
            "name": name,
            "group": group_for_student,
            "row": stripped_row,
            "email": stripped_row[1] if len(stripped_row) > 1 else "",
            "website": website_link,
        })

    return {
        "assignment_order": assignment_order,
        "assignment_columns": assignment_columns,
        "column_headers": column_headers,
        "group_counts": group_counts,
        "student_count": student_count,
        "students": students,
    }


def build_overview(path: Path, data: dict) -> str:
    lines: list[str] = [f"Bestand: {path.name}"]
    lines.append(f"Totale kolommen: {len(data['column_headers'])}")

    if data["assignment_order"]:
        lines.append("Opdrachten en kolommen:")
        for assignment in data["assignment_order"]:
            columns = data["assignment_columns"].get(assignment, [])
            column_names = ", ".join(name for _, name in columns) if columns else "-"
            lines.append(f"- {assignment}: {column_names}")
    else:
        lines.append("Geen opdrachtinformatie gevonden in de eerste rij.")

    if data["group_counts"]:
        lines.append("Leerlinggroepen:")
        for group_name, count in data["group_counts"]:
            lines.append(f"- {group_name}: {count} leerlingen")
    else:
        lines.append("Geen groepen of leerlingen gevonden.")

    lines.append(f"Totaal aantal leerlingen: {data['student_count']}")
    return "\n".join(lines)


def has_assignment_data(student_row: list[str], columns: list[tuple[int, str]]) -> bool:
    return any(idx < len(student_row) and bool(student_row[idx]) for idx, _ in columns)


def generate_assignment_pdf(
    path: Path,
    assignment: str,
    columns: list[tuple[int, str]],
    students: list[dict],
) -> None:
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=LEFT_MARGIN,
        rightMargin=RIGHT_MARGIN,
        topMargin=TOP_MARGIN,
        bottomMargin=BOTTOM_MARGIN,
    )

    students_with_data = [
        student for student in students if has_assignment_data(student["row"], columns)
    ]
    if not students_with_data:
        raise ValueError("Geen beoordelingsgegevens om in het PDF te plaatsen.")

    story: list = []
    assignment_heading = _paragraph_text(assignment)

    score_columns, comment_column = split_assignment_columns(columns)
    for index, student in enumerate(students_with_data):

        story.append(Paragraph(assignment_heading, STYLE_ASSIGNMENT))
        story.append(
            Paragraph(
                f"<b>Leerling:</b> {_paragraph_text(student['name'])}",
                STYLE_INFO,
            )
        )
        if student.get("group"):
            story.append(
                Paragraph(
                    f"<b>Groep:</b> {_paragraph_text(student['group'])}",
                    STYLE_INFO,
                )
            )

        contact_fields = [
            (label, value)
            for label, value in (
                ("E-mail", student.get("email")),
                ("Website", student.get("website")),
            )
            if value
        ]

        if contact_fields:
            story.append(Paragraph("Contact", STYLE_SECTION))
            for label, raw in contact_fields:
                story.append(
                    Paragraph(
                        f"<b>{escape(_format_text(label))}:</b> {_paragraph_text(raw)}",
                        STYLE_INFO,
                    )
                )

        story.append(Paragraph("Beoordeling", STYLE_SECTION))
        table_rows = [
            [
                Paragraph("<b>Categorie</b>", STYLE_TABLE_HEADER),
                Paragraph("<b>Waarde</b>", STYLE_TABLE_HEADER),
            ]
        ]
        target_columns = score_columns or columns
        for col_index, column_name in target_columns:
            label = column_name if column_name else f"Kolom {col_index + 1}"
            value = _safe_row_value(student["row"], col_index)
            table_rows.append(
                [
                    Paragraph(_paragraph_text(label), STYLE_TABLE_BODY),
                    Paragraph(_paragraph_text(value), STYLE_TABLE_BODY),
                ]
            )

        table = Table(
            table_rows,
            colWidths=[TABLE_FIRST_COLUMN_WIDTH, doc.width - TABLE_FIRST_COLUMN_WIDTH],
            repeatRows=1,
            hAlign="LEFT",
        )
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 16))

        if comment_column:
            col_index, label = comment_column
            comment_value = _safe_row_value(student["row"], col_index)
            if comment_value:
                story.append(Paragraph(escape(_format_text(label)), STYLE_SECTION))
                story.append(Paragraph(_paragraph_text(comment_value), STYLE_INFO))
                story.append(Spacer(1, 16))

        if index < len(students_with_data) - 1:
            story.append(PageBreak())

    doc.build(story)


def main() -> None:
    root = tk.Tk()
    root.title("Excel-bestand kiezen")
    root.geometry("680x580")

    selected_var = tk.StringVar(value="Geen bestand geselecteerd")
    assignment_choice = tk.StringVar(value="")
    assignment_info_var = tk.StringVar(value="Geen opdracht geselecteerd.")
    students_selection_info = tk.StringVar(value="Geen leerlingen geselecteerd.")
    select_all_var = tk.BooleanVar(value=False)

    current_data: dict | None = None
    student_entries: list[dict] = []
    all_students: list[dict] = []

    def current_assignment_columns() -> list[tuple[int, str]]:
        if current_data is None:
            return []
        assignment = assignment_choice.get()
        if not assignment:
            return []
        return current_data["assignment_columns"].get(assignment, [])

    def count_with_data(students_subset: list[dict]) -> int:
        columns = current_assignment_columns()
        if not columns:
            return len(students_subset)
        return sum(1 for student in students_subset if has_assignment_data(student["row"], columns))

    def refresh_student_list(reset_select: bool = False) -> None:
        nonlocal student_entries
        columns = current_assignment_columns()
        if columns:
            filtered = [student for student in all_students if has_assignment_data(student["row"], columns)]
        else:
            filtered = list(all_students)

        student_entries = filtered
        student_listbox.config(state="normal")
        student_listbox.delete(0, tk.END)

        if not student_entries:
            if reset_select:
                select_all_var.set(False)
            student_listbox.config(state="disabled")
            students_selection_info.set("Geen leerlingen met resultaten gevonden.")
            return

        for student in student_entries:
            group = student["group"]
            display_group = group if group and group != "(onbekende groep)" else "Leerling"
            display_text = f"{display_group} – {student['name']}"
            student_listbox.insert(tk.END, display_text)

        if reset_select:
            select_all_var.set(False)

        if select_all_var.get():
            student_listbox.selection_clear(0, tk.END)
            student_listbox.config(state="disabled")
            total = len(student_entries)
            with_data = count_with_data(student_entries)
            students_selection_info.set(
                f"Alle leerlingen geselecteerd ({total} totaal, {with_data} met resultaten)."
            )
        else:
            student_listbox.config(state="normal")
            student_listbox.selection_clear(0, tk.END)
            update_selected_students_label()

    def update_summary(text: str) -> None:
        summary_box.config(state="normal")
        summary_box.delete("1.0", tk.END)
        summary_box.insert(tk.END, text)
        summary_box.config(state="disabled")

    def on_assignment_selected(_event: object | None = None) -> None:
        if current_data is None:
            assignment_info_var.set("Geen gegevens geladen.")
            return

        selection = assignment_choice.get()
        if not selection:
            assignment_info_var.set("Geen opdracht geselecteerd.")
            return

        columns = current_data["assignment_columns"].get(selection, [])
        if columns:
            column_names = ", ".join(name for _, name in columns)
            assignment_info_var.set(f"Kolommen voor '{selection}': {column_names}")
        else:
            assignment_info_var.set(f"Geen kolommen gevonden voor '{selection}'.")
        refresh_student_list(reset_select=True)

    def update_selected_students_label() -> None:
        if select_all_var.get():
            if not student_entries:
                students_selection_info.set("Geen leerlingen met resultaten gevonden.")
                return
            total = len(student_entries)
            with_data = count_with_data(student_entries)
            students_selection_info.set(
                f"Alle leerlingen geselecteerd ({total} totaal, {with_data} met resultaten)."
            )
            return
        selection = student_listbox.curselection()
        available = len(student_entries)
        if not selection:
            if available:
                students_selection_info.set(f"Geen leerlingen geselecteerd ({available} beschikbaar).")
            else:
                students_selection_info.set("Geen leerlingen met resultaten gevonden.")
            return
        names = [student_entries[index]["name"] for index in selection]
        selected_students = [student_entries[index] for index in selection]
        with_data = count_with_data(selected_students)
        students_selection_info.set(
            f"Geselecteerd ({len(selection)} totaal, {with_data} met resultaten): {', '.join(names)}"
        )

    def on_student_selection(_event: object | None = None) -> None:
        if not select_all_var.get():
            update_selected_students_label()

    def toggle_select_all() -> None:
        if not student_entries:
            select_all_var.set(False)
            students_selection_info.set("Geen leerlingen met resultaten gevonden.")
            return
        if select_all_var.get():
            student_listbox.selection_clear(0, tk.END)
            student_listbox.config(state="disabled")
            total = len(student_entries)
            with_data = count_with_data(student_entries)
            students_selection_info.set(
                f"Alle leerlingen geselecteerd ({total} totaal, {with_data} met resultaten)."
            )
        else:
            student_listbox.config(state="normal")
            update_selected_students_label()

    def collect_students(require_email: bool = False) -> tuple[str, list[tuple[int, str]], list[tuple[int, str]], tuple[int, str] | None, list[dict]] | None:
        if current_data is None:
            messagebox.showwarning("Geen gegevens", "Selecteer eerst een Excel-bestand.", parent=root)
            return None

        assignment = assignment_choice.get()
        if not assignment:
            messagebox.showwarning("Geen opdracht", "Kies eerst een opdracht.", parent=root)
            return None

        columns = current_data["assignment_columns"].get(assignment, [])
        if not columns:
            messagebox.showwarning("Geen kolommen", "Geen kolommen gevonden voor deze opdracht.", parent=root)
            return None

        if select_all_var.get():
            selected_students_local = student_entries
        else:
            indices = student_listbox.curselection()
            if not indices:
                messagebox.showwarning("Geen leerlingen", "Selecteer een of meer leerlingen.", parent=root)
                return None
            selected_students_local = [student_entries[index] for index in indices]

        students_with_data = [
            student for student in selected_students_local if has_assignment_data(student["row"], columns)
        ]
        if not students_with_data:
            messagebox.showwarning(
                "Geen resultaten",
                "Voor de geselecteerde leerlingen zijn geen gegevens beschikbaar bij deze opdracht.",
                parent=root,
            )
            return None

        score_columns, comment_column = split_assignment_columns(columns)

        if require_email:
            students_with_email = [student for student in students_with_data if student.get("email")]
            missing_email_names = [student["name"] for student in students_with_data if not student.get("email")]
            if not students_with_email:
                messagebox.showwarning(
                    "Geen e-mailadressen",
                    "Geen e-mailadressen gevonden voor de geselecteerde leerlingen.",
                    parent=root,
                )
                return None
            if missing_email_names:
                messagebox.showinfo(
                    "Leerlingen zonder e-mail",
                    "De volgende leerlingen worden overgeslagen omdat er geen e-mailadres beschikbaar is:\n"
                    + "\n".join(missing_email_names),
                    parent=root,
                )
            target_students = students_with_email
        else:
            target_students = students_with_data

        return assignment, columns, score_columns, comment_column, target_students

    def reset_state() -> None:
        nonlocal current_data, student_entries
        selected_var.set("Geen bestand geselecteerd")
        update_summary("")
        assignment_choice.set("")
        assignment_box.config(values=(), state="disabled")
        assignment_info_var.set("Geen opdracht geselecteerd.")
        select_all_var.set(False)
        student_listbox.config(state="disabled")
        student_listbox.delete(0, tk.END)
        student_entries = []
        students_selection_info.set("Geen leerlingen geselecteerd.")
        pdf_button.config(state="disabled")
        email_button.config(state="disabled")
        current_data = None

    def populate_students(data: dict) -> None:
        nonlocal all_students, student_entries
        all_students = data.get("students", [])
        if not all_students:
            student_entries = []
            student_listbox.delete(0, tk.END)
            student_listbox.config(state="disabled")
            students_selection_info.set("Geen leerlingen gevonden.")
            return
        refresh_student_list(reset_select=True)

    def choose_file() -> None:
        nonlocal current_data
        dialog_kwargs: dict[str, Any] = {
            "title": "Kies een Excel-bestand",
            "filetypes": [("Excel-bestanden", "*.xlsx *.xls"), ("Alle bestanden", "*.*")],
            "parent": root,
        }
        if DEFAULT_INITIAL_DIR.exists():
            dialog_kwargs["initialdir"] = str(DEFAULT_INITIAL_DIR)

        filepath = filedialog.askopenfilename(**dialog_kwargs)
        if not filepath:
            reset_state()
            return

        path = Path(filepath)
        selected_var.set(str(path))
        try:
            current_data = parse_excel(path)
            summary = build_overview(path, current_data)
        except RuntimeError as exc:
            messagebox.showerror("Fout bij lezen", str(exc), parent=root)
            reset_state()
            return
        except Exception as exc:
            messagebox.showerror("Onverwachte fout", str(exc), parent=root)
            reset_state()
            return

        update_summary(summary)
        populate_students(current_data)
        select_all_var.set(False)
        toggle_select_all()
        assignments = current_data["assignment_order"]
        if assignments:
            assignment_box.config(values=assignments, state="readonly")
            assignment_choice.set(assignments[0])
            assignment_box.current(0)
            on_assignment_selected()
        else:
            assignment_box.config(values=(), state="disabled")
            assignment_choice.set("")
            assignment_info_var.set("Geen opdrachten gevonden.")
        pdf_button.config(state="normal")
        email_button.config(state="normal")

    def _open_with_default_app(path: Path) -> None:
        # Open a file with the default OS handler (Windows/macOS/Linux)
        if sys.platform.startswith("win") or os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform.startswith("darwin"):
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])

    def run_pdf_export() -> None:
        result = collect_students(require_email=False)
        if result is None:
            return
        assignment, columns, _score_columns, _comment_column, students_with_data = result

        preview_path = Path.home() / "Downloads" / f"preview_{assignment}.pdf"
        try:
            generate_assignment_pdf(preview_path, assignment, columns, students_with_data)
        except Exception as exc:
            messagebox.showerror("Fout bij PDF-preview", str(exc), parent=root)
            return

        try:
            _open_with_default_app(preview_path)
        except Exception as exc:
            messagebox.showwarning(
                "Preview openen mislukt",
                f"De preview is wel opgeslagen ({preview_path}), maar kon niet automatisch worden geopend.\n"
                f"Foutmelding: {exc}",
                parent=root,
            )

        save_path = filedialog.asksaveasfilename(
            title="Sla PDF op",
            defaultextension=".pdf",
            filetypes=[("PDF-bestanden", "*.pdf"), ("Alle bestanden", "*.*")],
            parent=root,
            initialdir=str(Path.home() / "Downloads"),
            initialfile=f"{assignment}.pdf",
        )
        if not save_path:
            return

        try:
            generate_assignment_pdf(Path(save_path), assignment, columns, students_with_data)
        except Exception as exc:
            messagebox.showerror("Fout bij PDF", str(exc), parent=root)
            return

        # Verwijder de preview zodra de definitieve PDF is opgeslagen
        try:
            if preview_path.exists():
                preview_path.unlink()
        except Exception:
            # Als verwijderen faalt (bestand nog geopend), laat het staan zonder te storen
            pass

        messagebox.showinfo("Gereed", f"PDF opgeslagen als:\n{save_path}", parent=root)

    def run_email_flow() -> None:
        result = collect_students(require_email=True)
        if result is None:
            return
        assignment, _columns, score_columns, comment_column, students_with_email = result

        email_previews: list[tuple[str, str, str, str]] = []
        for student in students_with_email:
            recipient = student.get("email")
            if not recipient:
                continue
            subject = f"Resultaten {assignment}"
            body = build_email_body(assignment, student, score_columns, comment_column)
            preview_text = f"Onderwerp: {subject}\nAan: {recipient}\n\n{body}"
            email_previews.append((recipient, subject, body, preview_text))

        if not email_previews:
            messagebox.showwarning("Geen e-mailadressen", "Geen geldige e-mailadressen gevonden.", parent=root)
            return
        recipients_list = [recipient for recipient, _, _, _ in email_previews]
        preview_text = "\n---\n".join(preview for _, _, _, preview in email_previews)

        preview_window = tk.Toplevel(root)
        preview_window.title("E-mail preview")
        preview_window.geometry("720x520")
        preview_window.transient(root)
        preview_window.grab_set()

        preview_label = tk.Label(preview_window, text="Controleer het bericht voordat je het verzendt:")
        preview_label.pack(pady=(10, 5))

        preview_textbox = tk.Text(preview_window, wrap="word")
        preview_textbox.insert("1.0", preview_text)
        preview_textbox.config(state="disabled")
        preview_textbox.pack(fill="both", expand=True, padx=12, pady=5)

        button_frame = ttk.Frame(preview_window)
        button_frame.pack(pady=10)

        status_var = tk.StringVar(value="")
        status_label = tk.Label(preview_window, textvariable=status_var)
        status_label.pack(pady=(0, 8))

        sending_state: dict[str, bool] = {"active": False}

        def handle_failure(recipient: str, exc: Exception) -> None:
            sending_state["active"] = False
            status_var.set("")
            send_button.config(state="normal")
            cancel_button.config(state="normal")
            messagebox.showerror(
                "Verzenden mislukt",
                f"Fout bij verzenden naar {recipient}:\n{exc}",
                parent=preview_window,
            )

        def handle_success() -> None:
            sending_state["active"] = False
            preview_window.destroy()
            messagebox.showinfo(
                "E-mail verzonden",
                "E-mails verzonden naar:\n" + "\n".join(recipients_list),
                parent=root,
            )

        def worker() -> None:
            try:
                for recipient, subject, body, _ in email_previews:
                    send_email_via_gmail(subject, body, [recipient])
            except Exception as exc:
                root.after(0, lambda r=recipient, e=exc: handle_failure(r, e))
                return
            root.after(0, handle_success)

        def send_and_close() -> None:
            if sending_state["active"]:
                return
            sending_state["active"] = True
            send_button.config(state="disabled")
            cancel_button.config(state="disabled")
            status_var.set("Bezig met verzenden...")
            threading.Thread(target=worker, daemon=True).start()

        def cancel_preview() -> None:
            if sending_state["active"]:
                messagebox.showwarning(
                    "Bezig met verzenden",
                    "Wacht tot het verzenden is afgerond.",
                    parent=preview_window,
                )
                return
            preview_window.destroy()

        send_button = ttk.Button(button_frame, text="Verzenden", command=send_and_close)
        send_button.pack(side=tk.LEFT, padx=10)
        cancel_button = ttk.Button(button_frame, text="Annuleren", command=cancel_preview)
        cancel_button.pack(side=tk.LEFT, padx=10)

        preview_window.protocol("WM_DELETE_WINDOW", cancel_preview)

    instructions = tk.Label(root, text="Selecteer een Excel-bestand:", font=("Helvetica", 14))
    instructions.pack(pady=(20, 10))

    browse_button = tk.Button(root, text="Bladeren...", command=choose_file)
    browse_button.pack()

    selected_label = tk.Label(root, textvariable=selected_var, wraplength=640, justify="center")
    selected_label.pack(padx=20, pady=(15, 5))

    assignment_frame = ttk.LabelFrame(root, text="Opdracht kiezen")
    assignment_frame.pack(fill="x", padx=20, pady=(10, 10))

    assignment_box = ttk.Combobox(assignment_frame, textvariable=assignment_choice, state="disabled")
    assignment_box.pack(fill="x", padx=10, pady=(10, 5))
    assignment_box.bind("<<ComboboxSelected>>", on_assignment_selected)

    assignment_info_label = tk.Label(
        assignment_frame,
        textvariable=assignment_info_var,
        wraplength=600,
        justify="center",
    )
    assignment_info_label.pack(padx=10, pady=(0, 10))

    students_frame = ttk.LabelFrame(root, text="Leerlingen selecteren")
    students_frame.pack(fill="both", expand=False, padx=20, pady=(0, 10))

    select_all_check = ttk.Checkbutton(
        students_frame,
        text="Alle leerlingen",
        variable=select_all_var,
        command=toggle_select_all,
    )
    select_all_check.pack(anchor="w", padx=10, pady=(10, 5))

    student_listbox = tk.Listbox(
        students_frame,
        selectmode=tk.MULTIPLE,
        activestyle="dotbox",
        exportselection=False,
        height=10,
    )
    student_listbox.pack(fill="both", expand=True, padx=10, pady=(0, 5))
    student_listbox.bind("<<ListboxSelect>>", on_student_selection)
    student_listbox.config(state="disabled")

    students_info_label = tk.Label(
        students_frame,
        textvariable=students_selection_info,
        wraplength=600,
        justify="left",
    )
    students_info_label.pack(fill="x", padx=10, pady=(0, 10))

    buttons_frame = ttk.Frame(root)
    buttons_frame.pack(pady=(0, 10))

    pdf_button = ttk.Button(buttons_frame, text="PDF maken", command=run_pdf_export, state="disabled")
    pdf_button.pack(side=tk.LEFT, padx=8)

    email_button = ttk.Button(buttons_frame, text="E-mailen", command=run_email_flow, state="disabled")
    email_button.pack(side=tk.LEFT, padx=8)

    summary_box = tk.Text(root, wrap="word", height=8, state="disabled")
    summary_box.pack(fill="both", expand=True, padx=20, pady=(5, 20))

    root.mainloop()


if __name__ == "__main__":
    main()
